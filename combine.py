import requests
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage

def select_image():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def check_expected_text(extracted_text, expected_text):
    return expected_text in extracted_text

def resize_image(img, width, height):
    return img.resize((width, height))

def pil_to_bytes(img):
    img_bytes = BytesIO()
    img.save(img_bytes, format='PNG')
    return img_bytes.getvalue()

# Prompt user for number of images
num_images = int(input("Enter the number of images: "))

# Create a new Workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Initialize variables to track the current row
current_row = 1

# Loop through each image
for i in range(num_images):
    # Select image
    image_path = select_image()
    files = {'image': open(image_path, 'rb')}

    response = requests.post('http://127.0.0.1:5000/extract_text', files=files)

    if response.status_code == 200:
        extracted_text = response.json()['text']
        print("Extracted Text:")
        print(extracted_text)

        # Define the expected text
        expected_text = "50°NE"

        # Check if the expected text is present in the extracted text
        if check_expected_text(extracted_text, expected_text):
            print("Image contains expected text:", expected_text)
            # Ask user for text
            text = input("Enter text for this image: ")
            img = PILImage.open(image_path)
            desired_width = 150
            desired_height = 150
            img_resized = resize_image(img, desired_width, desired_height)
            img_bytes = pil_to_bytes(img_resized)
            img_obj = ExcelImage(BytesIO(img_bytes))
            img_obj.width, img_obj.height = desired_width, desired_height
            ws.add_image(img_obj, f'B{current_row}')
            cell_text = ws.cell(row=current_row, column=1, value=text)
            cell_text.alignment = Alignment(wrap_text=True)
            orange_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            for paragraph in text.split('\n'):
                start_idx = text.index(paragraph)
                end_idx = start_idx + len(paragraph)
                for j in range(start_idx, end_idx):
                    cell_text.font = Font(size=22, color="000000", bold=True)
                    cell_text.alignment = Alignment(vertical='center')
                    cell_text.fill = orange_fill
            ws.column_dimensions['B'].width = desired_width * 0.15
            ws.row_dimensions[current_row].height = max(desired_height * 0.75, 30)
            current_row += 2  # Move to the next row with a gap of one row
        else:
            print("Image doesn't contain the expected text:", expected_text)
    else:
        print('Error:', response.json()['error'])

# Save the workbook
wb.save("output.xlsx")