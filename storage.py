
'''

import firebase_admin
from firebase_admin import credentials, storage, firestore
from datetime import datetime  # Import the datetime module

# Initialize Firebase Admin SDK
cred = credentials.Certificate("telecom-tower-performance-1-firebase-adminsdk-76b3k-265f93b36b.json")
firebase_admin.initialize_app(cred, {'storageBucket': 'telecom-tower-performance-1.appspot.com'})

# Initialize Firestore client
db = firestore.client()

# Get a reference to the default storage bucket
bucket = storage.bucket()

# Ask user for the folder name
folder_name = input("Enter the name of the folder to store the file in: ")

# Generate the document name with the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Format: YYYY-MM-DD_HH-MM-SS
document_name = f"{folder_name}_{current_datetime}"  # Append current date and time to the folder name

# Path to the local Excel file you want to upload
local_excel_file_path = "C:\\Users\\Gaurav\\Documents\\Python Scripts\\Python Folder G\\output.xlsx"

# Define the destination path in Firebase Storage with the folder name
destination_blob_name = f"excelFiles/{folder_name}/{document_name}.xlsx"

# Upload the local file to Firebase Storage
blob = bucket.blob(destination_blob_name)
blob.upload_from_filename(local_excel_file_path)

# Get the URL of the uploaded file
uploaded_file_url = blob.public_url

print("File uploaded to Firebase Storage.")

# Create a dictionary with document data
document_data = {
    "name": document_name,
    "url": uploaded_file_url
}

# Set the document in Firestore with the provided name
db.collection("files").document(document_name).set(document_data)

print("File URL and document name saved to Firestore.")

'''

'''import json

import requests
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage
import firebase_admin
from firebase_admin import credentials, storage, firestore
from datetime import datetime  # Import the datetime module


def select_image():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path


def check_expected_text(extracted_text, expected_text):
    return expected_text in extracted_text


def resize_image(img, width, height):
    return img.resize((width, height))


def pil_to_bytes(img):
    img_bytes = BytesIO()
    img.save(img_bytes, format='PNG')
    return img_bytes.getvalue()


# Prompt user for number of images
num_images = int(input("Enter the number of images: "))

# Create a new Workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Initialize variables to track the current row
current_row = 1

# Loop through each image

for i in range(num_images):
    # Select image
    image_path = select_image()
    files = {'image': open(image_path, 'rb')}

    response = requests.post('http://127.0.0.1:5000/extract_text', files=files)

    if response.status_code == 200:
        extracted_text = response.json()['text']
        print("Extracted Text:")
        print(extracted_text)

        # Define the expected text
        expected_text = "50°NE"

        # Check if the expected text is present in the extracted text
        if check_expected_text(extracted_text, expected_text):
            print("Image contains expected text:", expected_text)
            # Ask user for text
            text = input("Enter text for this image: ")
            img = PILImage.open(image_path)
            desired_width = 150
            desired_height = 150
            img_resized = resize_image(img, desired_width, desired_height)
            img_bytes = pil_to_bytes(img_resized)
            img_obj = ExcelImage(BytesIO(img_bytes))
            img_obj.width, img_obj.height = desired_width, desired_height
            ws.add_image(img_obj, f'B{current_row}')
            cell_text = ws.cell(row=current_row, column=1, value=text)
            cell_text.alignment = Alignment(wrap_text=True)
            orange_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            for paragraph in text.split('\n'):
                start_idx = text.index(paragraph)
                end_idx = start_idx + len(paragraph)
                for j in range(start_idx, end_idx):
                    cell_text.font = Font(size=22, color="000000", bold=True)
                    cell_text.alignment = Alignment(vertical='center')
                    cell_text.fill = orange_fill
            ws.column_dimensions['B'].width = desired_width * 0.15
            ws.row_dimensions[current_row].height = max(desired_height * 0.75, 30)
            current_row += 2  # Move to the next row with a gap of one row
        else:
            print("Image doesn't contain the expected text:", expected_text)

        # Check if response contains JSON data
        try:
            error_message = response.json().get('error')
            print('Error:', error_message)
        except ValueError:
            print('Error: Unexpected response format')

# Save the workbook to a BytesIO object
excel_bytes = BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)  # Reset the pointer to the beginning of the BytesIO object


# Initialize Firebase Admin SDK
cred = credentials.Certificate("telecom-tower-performance-1-firebase-adminsdk-76b3k-265f93b36b.json")
firebase_admin.initialize_app(cred, {'storageBucket': 'telecom-tower-performance-1.appspot.com'})

# Initialize Firestore client
db = firestore.client()

# Get a reference to the default storage bucket
bucket = storage.bucket()

# Ask user for the folder name
folder_name = input("Enter the name of the folder to store the file in: ")

# Generate the document name with the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Format: YYYY-MM-DD_HH-MM-SS
document_name = f"{folder_name}_{current_datetime}"  # Append current date and time to the folder name

# Define the destination path in Firebase Storage with the folder name
destination_blob_name = f"excelFiles/{folder_name}/{document_name}.xlsx"



# Generate the document name with the entered folder name
current_datetime = folder_name

# Define the destination path in Firebase Storage with the folder name
destination_blob_name = f"excelFiles/{folder_name}/{document_name}.xlsx"


# Upload the Excel file to Firebase Storage
blob = bucket.blob(destination_blob_name)
blob.upload_from_file(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Get the URL of the uploaded file
uploaded_file_url = blob.public_url

print("File uploaded to Firebase Storage.")

# Create a dictionary with document data
document_data = {
    "name": document_name,
    "url": uploaded_file_url
}

# Set the document in Firestore with the provided name
db.collection("files").document(document_name).set(document_data)

print("File URL and document name saved to Firestore.")'''





import json
import requests
import tkinter as tk
from openpyxl import Workbook  # Corrected import
#from openpyxl import filedialog  # Corrected import
from tkinter import filedialog

from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage
import firebase_admin
from firebase_admin import credentials, storage, firestore
from datetime import datetime



def select_image():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path


def check_expected_text(extracted_text, expected_text):
    return expected_text in extracted_text


def resize_image(img, width, height):
    return img.resize((width, height))


def pil_to_bytes(img):
    img_bytes = BytesIO()
    img.save(img_bytes, format='PNG')
    return img_bytes.getvalue()


# Prompt user for number of images
num_images = int(input("Enter the number of images: "))

# Create a new Workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Initialize variables to track the current row
current_row = 1

# Loop through each image
for i in range(num_images):
    # Select image
    image_path = select_image()
    files = {'image': open(image_path, 'rb')}

    response = requests.post('http://127.0.0.1:5000/extract_text', files=files)

    if response.status_code == 200:
        extracted_text = response.json()['text']
        print("Extracted Text:")
        print(extracted_text)

        # Define the expected text
        expected_text = "50°NE"

        # Check if the expected text is present in the extracted text
        if check_expected_text(extracted_text, expected_text):
            print("Image contains expected text:", expected_text)
            # Ask user for text
            text = input("Enter text for this image: ")
            img = PILImage.open(image_path)
            desired_width = 150
            desired_height = 150
            img_resized = resize_image(img, desired_width, desired_height)
            img_bytes = pil_to_bytes(img_resized)
            img_obj = ExcelImage(BytesIO(img_bytes))
            img_obj.width, img_obj.height = desired_width, desired_height
            ws.add_image(img_obj, f'B{current_row}')
            cell_text = ws.cell(row=current_row, column=1, value=text)
            cell_text.alignment = Alignment(wrap_text=True)
            orange_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            for paragraph in text.split('\n'):
                start_idx = text.index(paragraph)
                end_idx = start_idx + len(paragraph)
                for j in range(start_idx, end_idx):
                    cell_text.font = Font(size=22, color="000000", bold=True)
                    cell_text.alignment = Alignment(vertical='center')
                    cell_text.fill = orange_fill
            ws.column_dimensions['B'].width = desired_width * 0.15
            ws.row_dimensions[current_row].height = max(desired_height * 0.75, 30)
            current_row += 2  # Move to the next row with a gap of one row
        else:
            print("Image doesn't contain the expected text:", expected_text)

        # Check if response contains JSON data
        try:
            error_message = response.json().get('error')
            print('Error:', error_message)
        except ValueError:
            print('Error: Unexpected response format')

# Save the workbook to a BytesIO object
excel_bytes = BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)  # Reset the pointer to the beginning of the BytesIO object


# Initialize Firebase Admin SDK
cred = credentials.Certificate("telecom-tower-performance-1-firebase-adminsdk-76b3k-265f93b36b.json")
firebase_admin.initialize_app(cred, {'storageBucket': 'telecom-tower-performance-1.appspot.com'})

# Initialize Firestore client
db = firestore.client()
# Get a reference to the default storage bucket
bucket = storage.bucket()

# Ask user for the folder name
folder_name = input("Enter the name of the folder to store the file in: ")

# Prompt user for today's date
today_date = input("Enter today's date (YYYY-MM-DD): ")

# Prompt user for today's time
today_time = input("Enter today's time (HH:MM:SS): ")

# Generate the document name with folder name, date, and time
document_name = f"{folder_name}"

# Define the destination path in Firebase Storage with the folder name
destination_blob_name = f"excelFiles/{folder_name}/{document_name}.xlsx"

# Upload the Excel file to Firebase Storage
blob = bucket.blob(destination_blob_name)
blob.upload_from_file(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Get the URL of the uploaded file
uploaded_file_url = blob.public_url

print("File uploaded to Firebase Storage.")

# Create a dictionary with document data
document_data = {
    "name": document_name,
    "date": today_date,
    "time": today_time,
    "url": uploaded_file_url
}

# Set the document in Firestore with the provided name
db.collection("files").document(document_name).set(document_data)

print("File URL and document name saved to Firestore.")
